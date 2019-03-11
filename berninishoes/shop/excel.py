# -*- coding: utf-8 -*-

from openpyxl import Workbook
import json

def creaExcelPedido(pedido):
    
    wb = Workbook()
    
    if pedido.detalle != '':
        dfact = json.loads(pedido.detalle)['datoscliente']
        dprods = json.loads(pedido.detalle)['detalle']
        
        ws = wb.active
        ws['A1']="Resumen Pedido"
        ws['B1']="Zapatos Bernini"
        ws['C1']=pedido.fecha
        ws['A2']="Datos del cliente"
        ws['A3']="Nombre"
        ws['B3']=dfact['nombre']+' '+dfact['apellidos']
        ws['A4']="Direccion"
        ws['B4']=dfact['direccion']+', '+dfact['localidad']+', '+dfact['provincia']+', '+dfact['pais']
        ws['A5']="Telefono"
        ws['B5']=dfact['telefono']
        ws['A7']="Detalle"
        ws['A8']="Referencia"
        ws['B8']="Nombre"
        ws['C8']="Cantidad"
        ws['D8']="Precio (ud)"
        ws['E8']="Total"
        ws['F8']="Atributos"
        i = 9
        total=0
        for dprod in dprods:
            precio = 0
            cant = 0
            if dprod['precio']!='':
                precio = float(dprod['precio'])
            if dprod['cantidad']!='':
                cant = int(dprod['cantidad'])
            ws['A'+str(i)]=dprod["id"]
            ws['B'+str(i)]=dprod["nombre"]
            ws['C'+str(i)]=dprod["cantidad"]
            ws['D'+str(i)]=dprod["precio"]
            ws['E'+str(i)]=cant*precio
            if dprod['detalle']:
                frase=''
                for a in dprod['detalle']:
                    frase=frase+' '+a['nombre']+':'+a['valor']
                ws['F'+str(i)]=frase
            
            i=i+1
            total=total+(precio*cant)
            
        ws['A'+str(i)]="Total pedido"
        ws['B'+str(i)]=total
        ws['C'+str(i)]="Euros"
        
        nombrefichero = 'Pedido_'+str(pedido.id)+'.xslx'
        try:
            wb.save(nombrefichero)
            return nombrefichero
        except Exception as e:
            print str(e)